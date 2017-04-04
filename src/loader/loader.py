

import numpy as np
import openpyxl as pyxl


def test_loader(fname=None):
    fname = '/home/mattias/Dropbox/CommonCode/Lifetime_fitting/src/test_folder/500_1e2_1e8.dat'
    data = np.genfromtxt(fname, names=['nxc', 'tau'], usecols=(8, 9))
    return data


def sinton2014(fname):
    '''
    This grabs the calculated data from sinton
    Extracts columns A-G of the raw data page.
    Can't extract columns H-I as it is not complete data and this would take
    some more work
    This outputs the data as a structured array, with the names as the column
    '''

    wb = pyxl.load_workbook(fname, read_only=True, data_only=True)
    # the command page was only added in the later versions
    if 'Command' in wb.get_sheet_names():
        data = _openpyxl_Sinton2014_ExtractRawDatadata(wb)

    # remove all the nan values from the data
    data = data[~np.isnan(data['Tau (sec)'])]
    data.dtype.names = ['Time in s', 'Photovoltage', 'Reference Voltage', 'PCD with baseline shift.', 'Ref (suns)', 'Delta Suns', 'Generation (pairs/s)', 'Conductivity increase', 'Apparent CD', 'nxc',
                        'tau', 'Tau In Range', '1/Tau in Range', '1/Tau Corrected', '1/Tau corrected in Range', '1/Tau Fit (auger corrected)', 'ni', 'ni2/tau', 'Implied Voc', 'Implied Suns', 'MCD for Ref Cell>0.6V']

    return data


def _openpyxl_Sinton2014_ExtractRawDatadata(wb):
    '''
        reads the raw and caculated data from the 'Calc' sheet of a  sinton WCT-120 spreadsheet.

        inputs:
            wb:
             instance of a openpylx workbook.
    '''

    # make sure the sheet is in the book
    assert 'Calc' in wb.get_sheet_names()

    # get the worksheet
    ws = wb.get_sheet_by_name('Calc')

    # get first section of data
    values1 = np.array([[i.value for i in j] for j in ws['A9':'I133']],
                       dtype=np.float64)
    headers1 = tuple(
        [[j.value for j in i] for i in ws['A8':'I8']][0])

    # get second section of data
    values2 = np.array([[i.value for i in j] for j in ws['O9':'Z133']],
                       dtype=np.float64)
    headers2 = tuple(
        [[j.value for j in i] for i in ws['O8':'Z8']][0])

    # form into one array with names
    values = np.hstack((values1, values2))
    headers = headers1 + headers2

    Out = values.view(dtype=list(
        zip(headers, ["float64"] * len(headers)))).copy()

    return Out


loadersandext = {
    'test_loader': '*nothing',
    'sinton2014': '*.xlsm'
}

loaders = {
    'test_loader': test_loader,
    'sinton2014': sinton2014
}


if __name__ == '__main__':
    test_loader(
        '/home/mattias/Dropbox/CommonCode/Lifetime_fitting/src/test_folder/500_1e2_1e8.dat')
